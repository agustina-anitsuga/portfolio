# -*- coding: utf-8 -*-
"""Excel "portfolio-*" sheet: the position-by-position detail."""

from openpyxl.utils import get_column_letter

from .excel_style import ExcelStyle
from .sheet_column import MONEY_FORMAT, PERCENT_FORMAT, SheetColumn

COLUMNS = [
    SheetColumn("key", "Ticker", 10),
    SheetColumn("name", "Nombre", 28),
    SheetColumn("sector", "Sector", 16),
    SheetColumn("units", "Unid. Cartera", 12, MONEY_FORMAT),
    SheetColumn("units_sold", "Unid. Vendidas", 12, MONEY_FORMAT),
    SheetColumn("price_source", "Fuente Precio", 13),
    # --- everything in ARS ---
    SheetColumn("avg_cost_ars", "Precio Prom. (ARS)", 15, MONEY_FORMAT),
    SheetColumn("price_ars", "Precio Actual (ARS)", 15, MONEY_FORMAT),
    SheetColumn("invested_ars", "Invertido (ARS)", 15, MONEY_FORMAT),
    SheetColumn("value_ars", "Valor Actual (ARS)", 15, MONEY_FORMAT),
    SheetColumn("pl_abs_ars", "P&L $ (ARS)", 13, MONEY_FORMAT),
    SheetColumn("pl_pct_ars", "P&L % (ARS)", 12, PERCENT_FORMAT),
    SheetColumn("cost_of_sales_ars", "Costo Ventas (ARS)", 15, MONEY_FORMAT),
    SheetColumn("income_from_sales_ars", "Ingreso Ventas (ARS)", 16, MONEY_FORMAT),
    SheetColumn("realized_abs_ars", "P&L Realiz. $ (ARS)", 15, MONEY_FORMAT),
    SheetColumn("realized_pct_ars", "P&L Realiz. % (ARS)", 14, PERCENT_FORMAT),
    # --- everything in USD ---
    SheetColumn("avg_cost_usd", "Precio Prom. (USD)", 15, MONEY_FORMAT),
    SheetColumn("price_usd", "Precio Actual (USD)", 15, MONEY_FORMAT),
    SheetColumn("invested_usd", "Invertido (USD)", 15, MONEY_FORMAT),
    SheetColumn("value_usd", "Valor Actual (USD)", 15, MONEY_FORMAT),
    SheetColumn("pl_abs_usd", "P&L $ (USD)", 13, MONEY_FORMAT),
    SheetColumn("pl_pct_usd", "P&L % (USD)", 12, PERCENT_FORMAT),
    SheetColumn("cost_of_sales_usd", "Costo Ventas (USD)", 15, MONEY_FORMAT),
    SheetColumn("income_from_sales_usd", "Ingreso Ventas (USD)", 16, MONEY_FORMAT),
    SheetColumn("realized_abs_usd", "P&L Realiz. $ (USD)", 15, MONEY_FORMAT),
    SheetColumn("realized_pct_usd", "P&L Realiz. % (USD)", 14, PERCENT_FORMAT),
]

SEMAPHORE_KEYS = ("pl_abs_ars", "pl_abs_usd", "pl_pct_ars", "pl_pct_usd",
                  "realized_abs_ars", "realized_abs_usd", "realized_pct_ars", "realized_pct_usd")


class PortfolioSheetWriter:
    """One row per position, with every metric in both currencies."""

    def __init__(self, workbook):
        self._workbook = workbook

    def write(self, sheet_name, rows):
        sheet = self._workbook.create_sheet(sheet_name)
        self._write_header(sheet)
        self._write_rows(sheet, rows)
        self._decorate(sheet, len(rows))
        return sheet

    def _write_header(self, sheet):
        sheet.append([c.header for c in COLUMNS])
        for index in range(1, len(COLUMNS) + 1):
            ExcelStyle.header(sheet.cell(row=1, column=index), wrap=True)

    def _write_rows(self, sheet, rows):
        for row_index, row in enumerate(rows, start=2):
            for col_index, column in enumerate(COLUMNS, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=column.cell_value(row))
                if column.number_format:
                    cell.number_format = column.number_format
                cell.border = ExcelStyle.BORDER

    def _decorate(self, sheet, row_count):
        for index, column in enumerate(COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = column.width
        last_row = row_count + 1
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last_row, 1)}"
        for key in SEMAPHORE_KEYS:
            letter = get_column_letter([c.key for c in COLUMNS].index(key) + 1)
            ExcelStyle.pl_semaphore(sheet, f"{letter}2:{letter}{max(last_row, 2)}")
