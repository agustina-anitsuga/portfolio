# -*- coding: utf-8 -*-
"""Formatting shared by the Excel sheets."""

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelStyle:
    """Dark headers, thin borders and a green/red traffic light for P&L."""

    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    GREEN_FONT = Font(color="006100")
    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    RED_FONT = Font(color="9C0006")
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TITLE_FONT = Font(size=16, bold=True)
    BOLD_FONT = Font(bold=True)

    @classmethod
    def header(cls, cell, wrap=False):
        cell.fill = cls.HEADER_FILL
        cell.font = cls.HEADER_FONT
        if wrap:
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    @classmethod
    def pl_semaphore(cls, worksheet, cell_range):
        worksheet.conditional_formatting.add(cell_range, CellIsRule(
            operator="greaterThanOrEqual", formula=["0"], fill=cls.GREEN_FILL, font=cls.GREEN_FONT))
        worksheet.conditional_formatting.add(cell_range, CellIsRule(
            operator="lessThan", formula=["0"], fill=cls.RED_FILL, font=cls.RED_FONT))
