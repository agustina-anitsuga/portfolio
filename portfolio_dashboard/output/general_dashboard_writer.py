# -*- coding: utf-8 -*-
"""Excel "dashboard-general" sheet: the whole portfolio together."""

from openpyxl.utils import get_column_letter

from ..market import Market
from .excel_style import ExcelStyle
from .sheet_column import MONEY_FORMAT, PERCENT_FORMAT

HEADERS = ["Ticker", "Tipo", "Invertido (ARS)", "Valor Actual (ARS)", "P&L $ (ARS)", "P&L % (ARS)",
           "Invertido (USD)", "Valor Actual (USD)", "P&L $ (USD)", "P&L % (USD)"]
WIDTHS = [12, 12, 15, 15, 13, 12, 15, 15, 13, 12]
NUMBER_FORMATS = {3: MONEY_FORMAT, 4: MONEY_FORMAT, 5: MONEY_FORMAT, 6: PERCENT_FORMAT,
                  7: MONEY_FORMAT, 8: MONEY_FORMAT, 9: MONEY_FORMAT, 10: PERCENT_FORMAT}
SEMAPHORE_COLUMNS = (5, 6, 9, 10)


class GeneralDashboardWriter:
    """One row per product of any market, sorted by P&L."""

    def __init__(self, workbook):
        self._workbook = workbook

    def write(self, reports):
        sheet = self._workbook.create_sheet("dashboard-general")
        sheet["A1"] = "Dashboard General - Resultado por Producto"
        sheet["A1"].font = ExcelStyle.TITLE_FONT
        header_row = self._write_header(sheet)
        for row in self._sorted_rows(reports):
            sheet.append(self._values(row))
        self._decorate(sheet, header_row, sheet.max_row)
        return sheet

    @staticmethod
    def _write_header(sheet):
        sheet.append([])
        sheet.append(HEADERS)
        header_row = sheet.max_row
        for index in range(1, len(HEADERS) + 1):
            ExcelStyle.header(sheet.cell(row=header_row, column=index))
        return header_row

    @staticmethod
    def _sorted_rows(reports):
        rows = [row for report in reports.values() for row in report.rows]
        rows.sort(key=lambda r: (r["pl_abs_ars"] is None, -(r["pl_abs_ars"] or -1e18)))
        return rows

    @staticmethod
    def _values(row):
        return [
            row["key"], Market.get(row["market"]).label,
            row["invested_ars"], row["value_ars"], row["pl_abs_ars"], _fraction(row["pl_pct_ars"]),
            row["invested_usd"], row["value_usd"], row["pl_abs_usd"], _fraction(row["pl_pct_usd"]),
        ]

    @staticmethod
    def _decorate(sheet, header_row, last_row):
        for column, number_format in NUMBER_FORMATS.items():
            for row in range(header_row + 1, last_row + 1):
                sheet.cell(row=row, column=column).number_format = number_format
        sheet.auto_filter.ref = f"A{header_row}:J{last_row}"
        sheet.freeze_panes = f"A{header_row + 1}"
        for column in SEMAPHORE_COLUMNS:
            letter = get_column_letter(column)
            ExcelStyle.pl_semaphore(sheet, f"{letter}{header_row + 1}:{letter}{last_row}")
        for index, width in enumerate(WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width


def _fraction(percent):
    return percent / 100.0 if percent is not None else None
