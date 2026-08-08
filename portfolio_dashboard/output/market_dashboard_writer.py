# -*- coding: utf-8 -*-
"""Excel "dashboard-*" sheet: KPIs and chart of one market."""

from openpyxl.chart import BarChart, Reference

from ..market import CURRENCIES
from .excel_style import ExcelStyle
from .sheet_column import MONEY_FORMAT, PERCENT_FORMAT

KPI_ROW = 3
TOP_COLUMN = 6
TOP_POSITIONS = 12


class MarketDashboardWriter:
    """The totals of a market in both currencies, plus a chart of invested
    vs current value for the largest positions."""

    def __init__(self, workbook):
        self._workbook = workbook

    def write(self, sheet_name, report):
        sheet = self._workbook.create_sheet(sheet_name)
        sheet["A1"] = f"Dashboard - {sheet_name.replace('dashboard-', '').upper()}"
        sheet["A1"].font = ExcelStyle.TITLE_FONT
        for index, currency in enumerate(CURRENCIES):
            self._write_kpis(sheet, report.kpis(currency), currency, column=1 + index * 2)
        top_rows = self._top_rows(report)
        self._write_top_table(sheet, top_rows)
        self._add_chart(sheet, len(top_rows))
        self._set_widths(sheet)
        return sheet

    @staticmethod
    def _write_kpis(sheet, kpis, currency, column):
        label = currency.upper()
        entries = [
            (f"Invertido ({label})", kpis.invested, False),
            (f"Valor Actual ({label})", kpis.value, False),
            (f"P&L No Realizado ({label})", kpis.pl_abs, False),
            (f"P&L No Realizado % ({label})", kpis.pl_pct, True),
            (f"P&L Realizado ({label})", kpis.realized, False),
        ]
        for index, (title, value, is_pct) in enumerate(entries):
            row = KPI_ROW + index
            sheet.cell(row=row, column=column, value=title).font = ExcelStyle.BOLD_FONT
            cell = sheet.cell(row=row, column=column + 1,
                              value=(value / 100.0 if (is_pct and value is not None) else value))
            cell.number_format = PERCENT_FORMAT if is_pct else MONEY_FORMAT

    @staticmethod
    def _top_rows(report):
        rows = sorted(report.rows, key=lambda r: -(r["value_ars"] or r["value_usd"] or 0))
        return rows[:TOP_POSITIONS]

    @staticmethod
    def _write_top_table(sheet, top_rows):
        headers = ("Ticker", "Invertido (ARS)", "Valor Actual (ARS)")
        for offset, header in enumerate(headers):
            sheet.cell(row=KPI_ROW, column=TOP_COLUMN + offset, value=header)
        for index, row in enumerate(top_rows, start=KPI_ROW + 1):
            sheet.cell(row=index, column=TOP_COLUMN, value=row["key"])
            sheet.cell(row=index, column=TOP_COLUMN + 1, value=row["invested_ars"] or 0)
            sheet.cell(row=index, column=TOP_COLUMN + 2, value=row["value_ars"] or 0)

    @staticmethod
    def _add_chart(sheet, top_count):
        chart = BarChart()
        chart.title = "Invertido vs Valor Actual - ARS (top posiciones)"
        chart.type = "col"
        chart.style = 10
        last_row = KPI_ROW + top_count
        data = Reference(sheet, min_col=TOP_COLUMN + 1, max_col=TOP_COLUMN + 2,
                         min_row=KPI_ROW, max_row=last_row)
        categories = Reference(sheet, min_col=TOP_COLUMN, min_row=KPI_ROW + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width, chart.height = 20, 10
        sheet.add_chart(chart, "A11")

    @staticmethod
    def _set_widths(sheet):
        for letter, width in (("A", 30), ("B", 16), ("C", 30), ("D", 16)):
            sheet.column_dimensions[letter].width = width
