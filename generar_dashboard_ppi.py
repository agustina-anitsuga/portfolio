#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_dashboard_ppi.py

Lee "Portfolio - Transacciones.xlsx" (hojas: instrumentos, config, tx-usd,
tx-cedears, tx-merval, tx-rsu), consulta precios actuales via la API de PPI
(Portfolio Personal Inversiones) y genera "Portfolio Dashboard.html": un
dashboard interactivo con una pestana por tipo de instrumento (US Stocks,
Cedears, Acciones Merval, RSU) mas una general, cada una con:

  - precio promedio de compra, precio actual, dinero invertido, valor
    actual, P&L $ y P&L % -- TODO en pesos (ARS) y en dolares (USD) a la vez.
  - tratamiento de ventas: unidades vendidas, costo de ventas, ingreso de
    ventas, P&L realizado $ y %, tambien en ambas monedas.
  - buscador, filtro por sector, filtro por tipo de instrumento (Stock/ETF),
    filtro por anio de compra y filtro por moneda a mostrar, columnas
    ordenables por click, y graficos.

  El filtro de anio de compra esta en todas las pestanas y lista solo los
  anios en los que efectivamente hubo compras. Una posicion con compras en
  varios anios aparece en cada uno de esos anios, y se muestra completa (no
  se prorratea al tramo comprado ese anio).

  RSU se trata exactamente igual que los demas portfolios (misma logica
  de costo promedio, P&L, dual-moneda); solo cambia que la hoja de origen es
  "tx-rsu" en vez de "tx-usd".

COMO SE CONVIERTE ENTRE ARS Y USD
    Costo, invertido y ventas se calculan por OPERACION, usando el monto real
    que hayas cargado en cada moneda:
    - Cedears: tx-cedears siempre trae el monto real en ARS (@Local, lo que
      opera en BYMA) y en USD (@Origen) por cada operacion.
    - Merval: tx-merval trae "Total amount (ARS)" siempre, y "Total amount
      (USD)" si lo cargaste -- viene calculado con el tipo de cambio del DIA
      de esa operacion, tal como lo tenias en la planilla original.
    - USD: tx-usd trae "Total amount (@Origin, USD)" siempre, y "Total
      amount (ARS)" si lo cargaste. Las operaciones viejas nunca tuvieron
      este dato, asi que para esas no hay forma de saber el ARS historico.
    Si una operacion puntual no tiene el monto en la otra moneda cargado, esa
    UNA operacion (no toda la posicion) se convierte al tipo de cambio de
    HOY como aproximacion, y la posicion queda marcada en el dashboard para
    que sepas que una parte es aproximada.

    El precio actual y el valor de mercado son "de hoy" por definicion, asi
    que esos sí siempre usan el tipo de cambio actual en ambas monedas.

    El tipo de cambio de hoy se obtiene primero intentando el dolar MEP via
    PPI (AL30 en ARS / AL30D en USD, el mismo par que PPI usa de ejemplo
    para su feed en tiempo real). Si PPI no esta disponible, se usa el valor
    manual de la hoja "config".

REQUISITOS
    pip install openpyxl ppi-client yfinance

    yfinance es opcional: si no esta instalado, el script sigue funcionando
    igual, simplemente no usa el fallback de Yahoo Finance (ver mas abajo).

CREDENCIALES PPI
    1. Entra a tu cuenta de PPI -> Gestiones -> Gestion de servicio API -> activar.
    2. Vas a obtener una Public Key y una Private Key.
    3. Expone las credenciales como variables de entorno antes de correr el script:

        export PPI_PUBLIC_KEY="tu_public_key"
        export PPI_PRIVATE_KEY="tu_private_key"

    Orden de fallback para el precio de cada instrumento:
    1. PPI en vivo.
    2. Si es un instrumento en USD (US Stocks o RSU) y PPI no devolvio nada
       (sin credenciales, ticker que PPI no opera, o simplemente sin datos
       porque el mercado de EEUU esta cerrado ahora) -> Yahoo Finance, que
       no requiere API key y devuelve el ultimo precio disponible aunque el
       mercado este cerrado. Esto NO aplica a Cedears (el precio de la
       Cedear en BYMA no es el mismo que el de la accion subyacente en
       USD -- usar Yahoo ahi daria un numero incorrecto).
    3. Precio manual de la columna "Precio Manual" en la hoja instrumentos.
    4. "no disponible" si ninguna de las anteriores funciono.

USO
    python3 generar_dashboard_ppi.py "Portfolio - Transacciones.xlsx" --out-html "Portfolio Dashboard.html"
"""

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference

try:
    from ppi_client.ppi import PPI
    HAVE_PPI = True
except ImportError:
    HAVE_PPI = False

try:
    import yfinance as yf
    HAVE_YFINANCE = True
except ImportError:
    HAVE_YFINANCE = False

PPI_PUBLIC_KEY = os.environ.get("PPI_PUBLIC_KEY", "")
PPI_PRIVATE_KEY = os.environ.get("PPI_PRIVATE_KEY", "")
PPI_SANDBOX = os.environ.get("PPI_SANDBOX", "false").lower() == "true"

MARKETS = ["usd", "cedears", "merval", "rsu", "bonds"]
# native currency actually quoted/traded for each market
NATIVE_CURRENCY = {"usd": "usd", "cedears": "ars", "merval": "ars", "rsu": "usd", "bonds": "ars"}
MARKET_LABELS = {"usd": "US Stocks", "cedears": "Cedears", "merval": "Acciones Merval", "rsu": "RSU", "bonds": "Bonos"}


# ---------------------------------------------------------------------------
# PPI client (best-effort: any failure just means "no live price")
# ---------------------------------------------------------------------------

_ppi_client = None
_ppi_tried = False


def get_ppi_client():
    global _ppi_client, _ppi_tried
    if _ppi_tried:
        return _ppi_client
    _ppi_tried = True
    if not HAVE_PPI:
        print("Aviso: paquete 'ppi-client' no instalado (pip install ppi-client). "
              "Se usaran solo precios manuales.", file=sys.stderr)
        return None
    if not PPI_PUBLIC_KEY or not PPI_PRIVATE_KEY:
        print("Aviso: no hay credenciales PPI_PUBLIC_KEY/PPI_PRIVATE_KEY configuradas. "
              "Se usaran solo precios manuales.", file=sys.stderr)
        return None
    try:
        client = PPI(sandbox=PPI_SANDBOX)
        client.account.login_api(PPI_PUBLIC_KEY, PPI_PRIVATE_KEY)
        _ppi_client = client
        return client
    except Exception as e:
        print(f"Aviso: no se pudo autenticar con PPI ({e}). "
              "Se usaran solo precios manuales.", file=sys.stderr)
        return None


def fetch_ppi_price(ticker, tipo, settlement):
    client = get_ppi_client()
    if not client or not tipo or not settlement:
        return None
    try:
        data = client.marketdata.current(ticker, tipo, settlement)
        price = data.get("price") if isinstance(data, dict) else None
        if not price:
            return None
        price = float(price)
        # Los bonos (BONOS) cotizan en PPI "por cada 100 de valor nominal"
        # (convencion estandar de mercado), no por unidad -- si no se ajusta
        # esto, el valor de la posicion queda inflado ~100x. Ojo: esto NO
        # afecta el calculo del dolar MEP (AL30/AL30D), porque ahi se usa un
        # cociente entre dos precios BONOS y el factor 100 se cancela solo.
        if tipo == "BONOS":
            price = price / 100.0
        return price
    except Exception:
        return None


def fetch_yahoo_price(ticker):
    """Fallback para instrumentos en USD que PPI no cotiza o que estan sin
    precio porque el mercado de EEUU esta cerrado (PPI muchas veces solo
    tiene datos de instrumentos USD durante la rueda). Yahoo Finance no
    necesita API key y devuelve el ultimo precio disponible aunque el
    mercado este cerrado (no exige que sea "en vivo").

    Devuelve (precio, motivo_si_fallo). Prueba varias vias porque el
    formato de fast_info/info cambia bastante entre versiones de yfinance."""
    if not HAVE_YFINANCE:
        return None, "yfinance no esta instalado (pip install yfinance)"

    last_err = None

    try:
        fi = yf.Ticker(ticker).fast_info
        for k in ("last_price", "lastPrice", "regularMarketPrice"):
            try:
                v = fi[k]
            except Exception:
                v = getattr(fi, k, None)
            if v:
                return float(v), None
    except Exception as e:
        last_err = e

    try:
        info = yf.Ticker(ticker).info
        if isinstance(info, dict):
            for k in ("regularMarketPrice", "currentPrice", "previousClose"):
                v = info.get(k)
                if v:
                    return float(v), None
    except Exception as e:
        last_err = e

    try:
        hist = yf.Ticker(ticker).history(period="5d")
        if hist is not None and not hist.empty:
            closes = hist["Close"].dropna()
            if len(closes):
                return float(closes.iloc[-1]), None
    except Exception as e:
        last_err = e

    reason = f"Yahoo Finance no devolvio precio para '{ticker}'"
    if last_err is not None:
        reason += f" ({last_err})"
    return None, reason


def fetch_ppi_trend_30d(ticker, tipo, settlement):
    """Best-effort 30-day % change using historical data. Returns None on any issue."""
    client = get_ppi_client()
    if not client or not tipo or not settlement:
        return None
    try:
        date_to = dt.datetime.now()
        date_from = date_to - dt.timedelta(days=30)
        hist = client.marketdata.search(ticker, tipo, settlement, date_from, date_to)
        prices = [h["price"] for h in hist if h.get("price")]
        if len(prices) < 2:
            return None
        return (prices[-1] - prices[0]) / prices[0] * 100
    except Exception:
        return None


def fetch_fx_rate_ppi():
    """Dolar MEP implicito via PPI: precio en ARS de AL30 / precio en USD de
    AL30D (mismo bono, dos monedas de liquidacion distintas). Es el mismo par
    que PPI usa de ejemplo para su feed de tiempo real.

    Devuelve (valor, motivo_si_fallo). AL30/AL30D cotizan en BYMA, asi que
    fuera del horario de mercado argentino (aprox. 11 a 17hs, dias habiles)
    es normal que no devuelvan precio -- no es un problema del script."""
    client = get_ppi_client()
    if not client:
        return None, "sin cliente PPI (sin credenciales o paquete ppi-client no instalado)"
    try:
        ars = client.marketdata.current("AL30", "BONOS", "INMEDIATA")
        usd = client.marketdata.current("AL30D", "BONOS", "INMEDIATA")
        p_ars = ars.get("price") if isinstance(ars, dict) else None
        p_usd = usd.get("price") if isinstance(usd, dict) else None
        if p_ars and p_usd:
            return float(p_ars) / float(p_usd), None
        faltantes = [t for t, p in (("AL30", p_ars), ("AL30D", p_usd)) if not p]
        return None, (f"PPI no devolvio precio para {' y '.join(faltantes)} "
                       f"(probablemente fuera del horario de mercado)")
    except Exception as e:
        return None, f"error consultando PPI: {e}"


def get_fx_rate(manual_fx):
    live, reason = fetch_fx_rate_ppi()
    if live:
        return live, "PPI (MEP AL30/AL30D, en vivo)"
    if manual_fx:
        return float(manual_fx), "manual (hoja config)"
    return None, reason or "no disponible"


# ---------------------------------------------------------------------------
# Read input workbook
# ---------------------------------------------------------------------------

def load_instrumentos(wb):
    ws = wb["instrumentos"]
    meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if not key:
            continue
        meta[key] = {
            "name": row[1] or key,
            "tipo_ppi": row[2] or "",
            "settlement": row[3] or "",
            "ratio": row[4] or "",
            "sector": row[5] or "",
            "currency": row[6] or "",
            "manual_price": row[7],
            "instrument_type": _col(row, 8) or "",
        }
    return meta


def load_config(wb):
    if "config" not in wb.sheetnames:
        return {"manual_fx": None}
    ws = wb["config"]
    manual_fx = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and "usd/ars" in str(row[0]).lower():
            manual_fx = row[1]
    return {"manual_fx": manual_fx}


def _parse_date(v):
    """(anio, mes, dia) de una fecha de transaccion, o None si no se puede leer.

    Las hojas mezclan formatos segun como se cargo cada fila: celdas de fecha
    reales de Excel, ISO ('2026-11-06') y dd/mm/aaaa ('30/09/2025'), asi que
    hay que contemplar los tres. Cuando la fecha viene como texto y NO empieza
    con el anio, se asume dia primero (formato local), que es como estan
    cargadas las hojas.
    """
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.year, v.month, v.day)
    s = str(v).strip()
    if not s:
        return None
    s = s.split()[0]  # descarta la hora si viene pegada ('2025-01-15 00:00:00')
    for sep in ("/", "-", "."):
        s = s.replace(sep, " ")
    parts = s.split()
    if len(parts) < 3 or not all(p.isdigit() for p in parts[:3]):
        return None
    a, b, c = (int(p) for p in parts[:3])
    if len(parts[0]) == 4:
        year, month, day = a, b, c      # ISO: aaaa mm dd
    else:
        day, month, year = a, b, c      # local: dd mm aaaa
    if year < 100:
        return None                     # anio de 2 digitos: ambiguo, se descarta
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    return (year, month, day)


def _year_of(v):
    """Anio ('2024') de una fecha, o None. Se devuelve como string para que
    compare directo contra el value de un <select> en el HTML."""
    parsed = _parse_date(v)
    return str(parsed[0]) if parsed else None


def load_dual_positions(rows, native, fx_rate):
    """Average-cost method tracking BOTH currencies at once, per transaction.

    rows: iterable of (key, op, units, amount_native, amount_other_or_None, date).
    `native` is 'ars' or 'usd' -- the currency amount_native is denominated in.
    `date` solo se usa para registrar en que anios se compro la posicion
    (filtro "Anio de compra" del dashboard); no afecta ningun calculo.

    When a transaction has a real recorded amount in the other currency
    (e.g. tx-cedears always has both @Local and @Origin; tx-merval has both
    ARS and USD from the original sheet), that real historical amount is
    used -- no FX approximation involved. Only when a transaction is
    missing the other-currency amount (e.g. legacy tx-usd rows that were
    never recorded in ARS) does it fall back to converting that single
    transaction at TODAY's fx_rate, and the position gets flagged so the UI
    can say so.
    """
    other = "usd" if native == "ars" else "ars"
    positions = {}
    for key, op, units, amt_native, amt_other, date in rows:
        if not key or not op or units is None:
            continue
        pos = positions.setdefault(key, {
            "units": 0.0, "units_sold": 0.0,
            "cost_basis_ars": 0.0, "cost_basis_usd": 0.0,
            "cost_of_sales_ars": 0.0, "cost_of_sales_usd": 0.0,
            "income_from_sales_ars": 0.0, "income_from_sales_usd": 0.0,
            "approx_used": False, "buy_years": set(),
        })
        units = float(units)
        amt_native = float(amt_native) if amt_native is not None else 0.0
        if amt_other is not None:
            amt_other = float(amt_other)
        elif fx_rate:
            amt_other = amt_native * fx_rate if native == "usd" else amt_native / fx_rate
            pos["approx_used"] = True
        else:
            amt_other = 0.0
            pos["approx_used"] = True

        amounts = {native: amt_native, other: amt_other}

        if op == "BUY":
            pos["units"] += units
            pos["cost_basis_ars"] += amounts["ars"]
            pos["cost_basis_usd"] += amounts["usd"]
            year = _year_of(date)
            if year:
                pos["buy_years"].add(year)
        elif op == "SELL":
            avg_ars = pos["cost_basis_ars"] / pos["units"] if pos["units"] > 1e-9 else 0.0
            avg_usd = pos["cost_basis_usd"] / pos["units"] if pos["units"] > 1e-9 else 0.0
            cost_ars = avg_ars * units
            cost_usd = avg_usd * units
            pos["cost_basis_ars"] -= cost_ars
            pos["cost_basis_usd"] -= cost_usd
            pos["units"] -= units
            pos["units_sold"] += units
            pos["cost_of_sales_ars"] += cost_ars
            pos["cost_of_sales_usd"] += cost_usd
            pos["income_from_sales_ars"] += amounts["ars"]
            pos["income_from_sales_usd"] += amounts["usd"]

    out = {}
    for k, p in positions.items():
        p["dual_real"] = not p.pop("approx_used")
        p["buy_years"] = sorted(p["buy_years"])
        out[k] = p
    return out


def _col(row, idx):
    return row[idx] if idx < len(row) else None


def load_all_positions(wb, fx_rate):
    out = {}

    tws = wb["tx-usd"]
    tx_usd_rows = list(tws.iter_rows(min_row=2, values_only=True))
    # col5 = Total amount (@Origin, USD) -- always present.
    # col6 = Total amount (ARS), optional -- only exists for entries where
    # you recorded the real ARS you paid. Historic tx-usd rows never had
    # this, so those fall back to today's fx_rate (flagged in the UI).
    out["usd"] = load_dual_positions(
        ((r[0], r[1], r[2], r[5], _col(r, 6), _col(r, 3)) for r in tx_usd_rows),
        native="usd", fx_rate=fx_rate,
    )

    tws = wb["tx-cedears"]
    tx_ced_rows = list(tws.iter_rows(min_row=2, values_only=True))
    # Total amount (@Local, ARS) = col 4 -- what actually trades on BYMA.
    # Total amount (@Origin, USD) = col 6 -- real USD cost recorded per trade.
    # Using both real amounts (instead of ratio math) avoids the P&L bug we
    # hit earlier where multiplying USD price by local unit count without
    # dividing by Ratio inflated results ~60x.
    out["cedears"] = load_dual_positions(
        ((r[0], r[1], r[2], r[4], r[6], _col(r, 3)) for r in tx_ced_rows),
        native="ars", fx_rate=fx_rate,
    )

    tws = wb["tx-merval"]
    tx_mer_rows = list(tws.iter_rows(min_row=2, values_only=True))
    # col5 = Total amount (ARS) -- always present.
    # col6 = Total amount (USD), restored from the original sheet: it was
    # already computed historically (Total amount ARS / USD-ARS rate on the
    # day of the trade), so it's real, not an approximation.
    out["merval"] = load_dual_positions(
        ((r[0], r[1], r[2], r[5], _col(r, 6), _col(r, 3)) for r in tx_mer_rows),
        native="ars", fx_rate=fx_rate,
    )

    if "tx-rsu" in wb.sheetnames:
        tws = wb["tx-rsu"]
        tx_rsu_rows = list(tws.iter_rows(min_row=2, values_only=True))
        # mismo layout que tx-usd: col5 = Total amount (@Origin, USD) siempre
        # presente (costo base = FMV al momento del vesting), col6 = Total
        # amount (ARS) opcional.
        out["rsu"] = load_dual_positions(
            ((r[0], r[1], r[2], r[5], _col(r, 6), _col(r, 3)) for r in tx_rsu_rows),
            native="usd", fx_rate=fx_rate,
        )
    else:
        out["rsu"] = {}

    if "tx-bonds" in wb.sheetnames:
        tws = wb["tx-bonds"]
        tx_bonds_rows = list(tws.iter_rows(min_row=2, values_only=True))
        # mismo layout que tx-merval: col5 = Total amount (ARS) siempre
        # presente, col6 = Total amount (USD) real (bonos como AO28 cotizan
        # y liquidan en las dos monedas en BYMA, asi que ambos montos vienen
        # de la operacion real, no aproximados).
        out["bonds"] = load_dual_positions(
            ((r[0], r[1], r[2], r[5], _col(r, 6), _col(r, 3)) for r in tx_bonds_rows),
            native="ars", fx_rate=fx_rate,
        )
    else:
        out["bonds"] = {}

    return out


def _fmt_date(v):
    """Fecha SIEMPRE como YYYY-MM-DD. Es lo que hace que la columna "Fecha"
    ordene bien: la tabla ordena strings, y en ese formato el orden
    alfabetico coincide con el cronologico. Si la hoja trae dd/mm/aaaa y se
    dejara asi, mezclado con las filas en ISO, el orden salia mal."""
    if v is None:
        return ""
    parsed = _parse_date(v)
    if parsed is None:
        return str(v)  # formato no reconocido: se muestra tal cual vino
    return "%04d-%02d-%02d" % parsed


def collect_transactions(wb):
    """Flat list of every BUY/SELL row across the three tx-* sheets, for the
    'Transacciones' tab -- click a ticker anywhere and jump here filtered."""
    txs = []

    def add(sheet, market, idx_units, idx_date, idx_amt_ars, idx_amt_usd):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = _col(row, 0)
            op = _col(row, 1)
            if not key or not op:
                continue
            units = _col(row, idx_units)
            date = _col(row, idx_date)
            amt_ars = _col(row, idx_amt_ars)
            amt_usd = _col(row, idx_amt_usd)
            u = float(units) if units is not None else None
            amt_ars = float(amt_ars) if amt_ars is not None else None
            amt_usd = float(amt_usd) if amt_usd is not None else None
            txs.append({
                "ticker": key, "market": market, "op": op, "date": _fmt_date(date),
                "year": _year_of(date),
                "units": u, "amount_ars": amt_ars, "amount_usd": amt_usd,
                "price_ars": (amt_ars / u) if (amt_ars is not None and u) else None,
                "price_usd": (amt_usd / u) if (amt_usd is not None and u) else None,
            })

    add("tx-usd", "usd", idx_units=2, idx_date=3, idx_amt_ars=6, idx_amt_usd=5)
    add("tx-cedears", "cedears", idx_units=2, idx_date=3, idx_amt_ars=4, idx_amt_usd=6)
    add("tx-merval", "merval", idx_units=2, idx_date=3, idx_amt_ars=5, idx_amt_usd=6)
    if "tx-rsu" in wb.sheetnames:
        add("tx-rsu", "rsu", idx_units=2, idx_date=3, idx_amt_ars=6, idx_amt_usd=5)
    if "tx-bonds" in wb.sheetnames:
        add("tx-bonds", "bonds", idx_units=2, idx_date=3, idx_amt_ars=5, idx_amt_usd=6)

    txs.sort(key=lambda t: (t["date"] or "", t["ticker"]), reverse=True)
    return txs


# ---------------------------------------------------------------------------
# Build report rows
# ---------------------------------------------------------------------------

def resolve_price_dual(key, market, meta, fx_rate):
    """Returns (price_ars, price_usd, source, debug_note). The instrument's
    native currency (whatever PPI/manual actually quotes) is exact; the
    other currency is derived via fx_rate (today's rate). debug_note solo
    se completa cuando termina sin precio, para poder mostrar en consola
    por que fallo cada fuente que se intento."""
    m = meta.get(key, {})
    tipo = m.get("tipo_ppi")
    settlement = m.get("settlement")
    native = NATIVE_CURRENCY[market]

    debug_note = None
    live = fetch_ppi_price(key, tipo, settlement)
    yahoo = yahoo_err = None
    if live is None and native == "usd":
        yahoo, yahoo_err = fetch_yahoo_price(key)

    if live is not None:
        native_price, source = live, "PPI (en vivo)"
    elif yahoo is not None:
        # PPI no cotiza este ticker o esta sin datos ahora mismo (comun
        # fuera del horario de mercado de EEUU) -- Yahoo Finance como
        # segunda fuente, solo para instrumentos en USD.
        native_price, source = yahoo, "Yahoo Finance"
    elif m.get("manual_price") is not None:
        native_price, source = float(m["manual_price"]), "manual"
    else:
        native_price, source = None, "no disponible"
        if native == "usd":
            debug_note = yahoo_err or ("yfinance no esta instalado (pip install yfinance)"
                                        if not HAVE_YFINANCE else None)

    price_ars = price_usd = None
    if native == "ars":
        price_ars = native_price
        if native_price is not None and fx_rate:
            price_usd = native_price / fx_rate
    else:
        price_usd = native_price
        if native_price is not None and fx_rate:
            price_ars = native_price * fx_rate

    return price_ars, price_usd, source, debug_note


def build_market_report(market, positions, meta, fx_rate):
    rows = []
    for key, pos in positions.items():
        if pos["units"] <= 1e-9 and pos["units_sold"] <= 1e-9:
            continue
        m = meta.get(key, {})
        price_ars, price_usd, source, price_debug_note = resolve_price_dual(key, market, meta, fx_rate)
        trend = fetch_ppi_trend_30d(key, m.get("tipo_ppi"), m.get("settlement"))
        units = max(pos["units"], 0.0)

        def per_currency(cost_basis, cost_of_sales, income_from_sales, price):
            avg_cost = cost_basis / units if (cost_basis is not None and units > 1e-9) else None
            value = price * units if price is not None else None
            pl_abs = (value - cost_basis) if (value is not None and cost_basis is not None) else None
            pl_pct = (pl_abs / cost_basis * 100) if (pl_abs is not None and cost_basis) else None
            realized_abs = (
                (income_from_sales - cost_of_sales)
                if (income_from_sales is not None and cost_of_sales is not None) else None
            )
            realized_pct = (realized_abs / cost_of_sales * 100) if (realized_abs is not None and cost_of_sales) else None
            return {
                "avg_cost": avg_cost, "price": price, "invested": cost_basis, "value": value,
                "pl_abs": pl_abs, "pl_pct": pl_pct,
                "cost_of_sales": cost_of_sales, "income_from_sales": income_from_sales,
                "realized_abs": realized_abs, "realized_pct": realized_pct,
            }

        ars = per_currency(pos["cost_basis_ars"], pos["cost_of_sales_ars"], pos["income_from_sales_ars"], price_ars)
        usd = per_currency(pos["cost_basis_usd"], pos["cost_of_sales_usd"], pos["income_from_sales_usd"], price_usd)

        row = {
            "key": key, "market": market, "name": m.get("name", key),
            "sector": m.get("sector") or "-", "ratio": m.get("ratio") or "",
            "instrument_type": m.get("instrument_type") or "-",
            "units": units, "units_sold": pos["units_sold"], "trend_30d": trend,
            "buy_years": pos.get("buy_years", []),
            "price_source": source, "fx_approx": not pos["dual_real"],
            "native_currency": NATIVE_CURRENCY[market],
            "price_debug_note": price_debug_note,
        }
        for k, v in ars.items():
            row[f"{k}_ars"] = v
        for k, v in usd.items():
            row[f"{k}_usd"] = v
        rows.append(row)

    rows.sort(key=lambda r: (r["value_ars"] is None and r["value_usd"] is None,
                              -(r["value_ars"] or r["value_usd"] or 0)))

    # % que representa cada linea sobre el total del portfolio (misma
    # pestana), en cada moneda.
    total_value_ars = sum(r["value_ars"] for r in rows if r["value_ars"] is not None)
    total_value_usd = sum(r["value_usd"] for r in rows if r["value_usd"] is not None)
    for r in rows:
        r["pct_portfolio_ars"] = (
            r["value_ars"] / total_value_ars * 100
            if (r["value_ars"] is not None and total_value_ars) else None
        )
        r["pct_portfolio_usd"] = (
            r["value_usd"] / total_value_usd * 100
            if (r["value_usd"] is not None and total_value_usd) else None
        )
    return rows


def _build_price_lookup(reports):
    """ticker -> current price (both currencies), pulled from the same
    resolve_price_dual results already computed for the portfolio reports --
    no extra PPI calls needed."""
    lookup = {}
    for rows in reports.values():
        for r in rows:
            lookup[r["key"]] = {
                "current_price_ars": r.get("price_ars"),
                "current_price_usd": r.get("price_usd"),
                "current_price_source": r.get("price_source"),
            }
    return lookup


def build_all_reports(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    meta = load_instrumentos(wb)
    config = load_config(wb)
    fx_rate, fx_source = get_fx_rate(config.get("manual_fx"))
    positions = load_all_positions(wb, fx_rate)
    reports = {m: build_market_report(m, positions[m], meta, fx_rate) for m in MARKETS}
    transactions = collect_transactions(wb)

    price_lookup = _build_price_lookup(reports)
    for t in transactions:
        t.update(price_lookup.get(t["ticker"], {
            "current_price_ars": None, "current_price_usd": None, "current_price_source": "no disponible",
        }))
        # P&L % de esa operacion puntual: como vino el precio de esa
        # transaccion contra la cotizacion de hoy, en cada moneda.
        for currency in ("ars", "usd"):
            price = t.get(f"price_{currency}")
            current_price = t.get(f"current_price_{currency}")
            t[f"pl_pct_{currency}"] = (
                (current_price - price) / price * 100
                if (price and current_price is not None) else None
            )

    return reports, fx_rate, fx_source, transactions


# ---------------------------------------------------------------------------
# Excel writer helpers (optional secondary output, --out-xlsx)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PORTFOLIO_COLUMNS = [
    ("key", "Ticker", 10, None),
    ("name", "Nombre", 28, None),
    ("sector", "Sector", 16, None),
    ("units", "Unid. Cartera", 12, "#,##0.00"),
    ("units_sold", "Unid. Vendidas", 12, "#,##0.00"),
    ("price_source", "Fuente Precio", 13, None),
    # --- todo en ARS ---
    ("avg_cost_ars", "Precio Prom. (ARS)", 15, "#,##0.00"),
    ("price_ars", "Precio Actual (ARS)", 15, "#,##0.00"),
    ("invested_ars", "Invertido (ARS)", 15, "#,##0.00"),
    ("value_ars", "Valor Actual (ARS)", 15, "#,##0.00"),
    ("pl_abs_ars", "P&L $ (ARS)", 13, "#,##0.00"),
    ("pl_pct_ars", "P&L % (ARS)", 12, "0.00%"),
    ("cost_of_sales_ars", "Costo Ventas (ARS)", 15, "#,##0.00"),
    ("income_from_sales_ars", "Ingreso Ventas (ARS)", 16, "#,##0.00"),
    ("realized_abs_ars", "P&L Realiz. $ (ARS)", 15, "#,##0.00"),
    ("realized_pct_ars", "P&L Realiz. % (ARS)", 14, "0.00%"),
    # --- todo en USD ---
    ("avg_cost_usd", "Precio Prom. (USD)", 15, "#,##0.00"),
    ("price_usd", "Precio Actual (USD)", 15, "#,##0.00"),
    ("invested_usd", "Invertido (USD)", 15, "#,##0.00"),
    ("value_usd", "Valor Actual (USD)", 15, "#,##0.00"),
    ("pl_abs_usd", "P&L $ (USD)", 13, "#,##0.00"),
    ("pl_pct_usd", "P&L % (USD)", 12, "0.00%"),
    ("cost_of_sales_usd", "Costo Ventas (USD)", 15, "#,##0.00"),
    ("income_from_sales_usd", "Ingreso Ventas (USD)", 16, "#,##0.00"),
    ("realized_abs_usd", "P&L Realiz. $ (USD)", 15, "#,##0.00"),
    ("realized_pct_usd", "P&L Realiz. % (USD)", 14, "0.00%"),
]

PCT_KEYS = {"pl_pct_ars", "pl_pct_usd", "realized_pct_ars", "realized_pct_usd"}


def write_portfolio_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    headers = [c[1] for c in PORTFOLIO_COLUMNS]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _, width, fmt) in enumerate(PORTFOLIO_COLUMNS, start=1):
            val = row.get(key)
            if key in PCT_KEYS and val is not None:
                val = val / 100.0
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fmt:
                cell.number_format = fmt
            cell.border = BORDER

    for i, (_, _, width, _) in enumerate(PORTFOLIO_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    last_row = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(last_row,1)}"

    for col_key in ("pl_abs_ars", "pl_abs_usd", "pl_pct_ars", "pl_pct_usd",
                     "realized_abs_ars", "realized_abs_usd", "realized_pct_ars", "realized_pct_usd"):
        col_idx = [c[0] for c in PORTFOLIO_COLUMNS].index(col_key) + 1
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{max(last_row,2)}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT)
        )
    return ws


def _sum_or_none(rows, key):
    vals = [r[key] for r in rows if r.get(key) is not None]
    return sum(vals) if vals else None


def write_dashboard_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"Dashboard - {sheet_name.replace('dashboard-', '').upper()}"
    ws["A1"].font = Font(size=16, bold=True)

    for col_i, currency in enumerate(("ars", "usd")):
        invested = _sum_or_none(rows, f"invested_{currency}")
        value = _sum_or_none(rows, f"value_{currency}")
        pl_abs = (value - invested) if (value is not None and invested is not None) else None
        pl_pct = (pl_abs / invested * 100) if (pl_abs is not None and invested) else None
        realized = _sum_or_none(rows, f"realized_abs_{currency}")
        kpis = [
            (f"Invertido ({currency.upper()})", invested, False),
            (f"Valor Actual ({currency.upper()})", value, False),
            (f"P&L No Realizado ({currency.upper()})", pl_abs, False),
            (f"P&L No Realizado % ({currency.upper()})", pl_pct, True),
            (f"P&L Realizado ({currency.upper()})", realized, False),
        ]
        base_col = 1 + col_i * 2
        for i, (label, val, is_pct) in enumerate(kpis):
            row = 3 + i
            ws.cell(row=row, column=base_col, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=base_col + 1, value=(val / 100.0 if (is_pct and val is not None) else val))
            cell.number_format = "0.00%" if is_pct else "#,##0.00"

    start_col = 6
    ws.cell(row=3, column=start_col, value="Ticker")
    ws.cell(row=3, column=start_col + 1, value="Invertido (ARS)")
    ws.cell(row=3, column=start_col + 2, value="Valor Actual (ARS)")
    top_rows = sorted(rows, key=lambda r: -(r.get("value_ars") or r.get("value_usd") or 0))[:12]
    for i, r in enumerate(top_rows, start=4):
        ws.cell(row=i, column=start_col, value=r["key"])
        ws.cell(row=i, column=start_col + 1, value=r.get("invested_ars") or 0)
        ws.cell(row=i, column=start_col + 2, value=r.get("value_ars") or 0)

    chart = BarChart()
    chart.title = "Invertido vs Valor Actual - ARS (top posiciones)"
    chart.type = "col"
    chart.style = 10
    data = Reference(ws, min_col=start_col + 1, max_col=start_col + 2, min_row=3, max_row=3 + len(top_rows))
    cats = Reference(ws, min_col=start_col, min_row=4, max_row=3 + len(top_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 20, 10
    ws.add_chart(chart, "A11")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    return ws


def write_general_dashboard(wb, reports):
    ws = wb.create_sheet("dashboard-general")
    ws["A1"] = "Dashboard General - Resultado por Producto"
    ws["A1"].font = Font(size=16, bold=True)

    headers = ["Ticker", "Tipo", "Invertido (ARS)", "Valor Actual (ARS)", "P&L $ (ARS)", "P&L % (ARS)",
               "Invertido (USD)", "Valor Actual (USD)", "P&L $ (USD)", "P&L % (USD)"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    all_rows = [r for rows in reports.values() for r in rows]
    all_rows.sort(key=lambda r: (r.get("pl_abs_ars") is None, -(r.get("pl_abs_ars") or -1e18)))

    for r in all_rows:
        ws.append([
            r["key"], MARKET_LABELS[r["market"]],
            r.get("invested_ars"), r.get("value_ars"), r.get("pl_abs_ars"),
            (r["pl_pct_ars"] / 100.0) if r.get("pl_pct_ars") is not None else None,
            r.get("invested_usd"), r.get("value_usd"), r.get("pl_abs_usd"),
            (r["pl_pct_usd"] / 100.0) if r.get("pl_pct_usd") is not None else None,
        ])
    last_row = ws.max_row
    for col, fmt in [(3, "#,##0.00"), (4, "#,##0.00"), (5, "#,##0.00"), (6, "0.00%"),
                      (7, "#,##0.00"), (8, "#,##0.00"), (9, "#,##0.00"), (10, "0.00%")]:
        for row in range(header_row + 1, last_row + 1):
            ws.cell(row=row, column=col).number_format = fmt

    ws.auto_filter.ref = f"A{header_row}:J{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    for col_idx in (5, 6, 9, 10):
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}{header_row+1}:{col_letter}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT))

    widths = [12, 12, 15, 15, 13, 12, 15, 15, 13, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


# ---------------------------------------------------------------------------
# HTML dashboard (interactive: tabs, search, sector/type filters, sortable
# columns, dual-currency, sales treatment)
# ---------------------------------------------------------------------------

def _market_kpis(rows):
    kpis = {}
    for currency in ("ars", "usd"):
        invested = _sum_or_none(rows, f"invested_{currency}")
        value = _sum_or_none(rows, f"value_{currency}")
        pl_abs = (value - invested) if (value is not None and invested is not None) else None
        pl_pct = (pl_abs / invested * 100) if (pl_abs is not None and invested) else None
        realized = _sum_or_none(rows, f"realized_abs_{currency}")
        kpis[currency] = {"invested": invested, "value": value, "pl_abs": pl_abs,
                           "pl_pct": pl_pct, "realized": realized}
    kpis["unpriced"] = sum(1 for r in rows if r.get("value_ars") is None and r.get("value_usd") is None)
    return kpis


def render_html(reports, fx_rate, fx_source, transactions, out_path):
    payload = {"markets": {}, "general": [], "transactions": transactions,
               "fx_rate": fx_rate, "fx_source": fx_source}
    for market, rows in reports.items():
        payload["markets"][market] = {
            "label": MARKET_LABELS[market],
            "kpis": _market_kpis(rows),
            "rows": rows,
        }
        payload["general"].extend(rows)

    payload["general"].sort(key=lambda r: (r.get("pl_abs_ars") is None, -(r.get("pl_abs_ars") or -1e18)))
    payload["generated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    html = HTML_PAGE.replace("__DATA_JSON__", json.dumps(payload, ensure_ascii=False))
    Path(out_path).write_text(html, encoding="utf-8")


HTML_PAGE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Portfolio Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  :root {
    --bg:#0f1115; --card:#171a21; --border:#262b36; --text:#e7e9ee; --muted:#9aa2b1;
    --green:#3ddc84; --red:#ff5c5c; --accent:#6ea8fe; --hover:#1d212b;
  }
  * { box-sizing:border-box; }
  body { margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif;
         background:var(--bg); color:var(--text); padding:28px 32px 60px; }
  h1 { font-size:22px; margin:0 0 4px; }
  .sub { color:var(--muted); font-size:13px; margin-bottom:8px; }
  .fxbar { color:var(--muted); font-size:12px; margin-bottom:20px; background:var(--card);
           border:1px solid var(--border); border-radius:8px; padding:8px 14px; display:inline-block; }
  .tabs { display:flex; gap:6px; margin-bottom:20px; border-bottom:1px solid var(--border); }
  .tab { padding:10px 18px; cursor:pointer; color:var(--muted); font-size:14px; border-bottom:2px solid transparent; }
  .tab.active { color:var(--text); border-bottom-color:var(--accent); font-weight:600; }
  .page { display:none; }
  .page.active { display:block; }
  .kpi-title { font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:.05em; margin:0 0 8px; }
  .kpis { display:flex; gap:12px; flex-wrap:wrap; margin-bottom:16px; }
  .kpi { background:var(--card); border:1px solid var(--border); border-radius:12px; padding:12px 16px; min-width:130px; flex:1; }
  .kpi .label { color:var(--muted); font-size:10.5px; text-transform:uppercase; letter-spacing:.04em; }
  .kpi .val { font-size:18px; font-weight:600; margin-top:4px; }
  .pos { color:var(--green); } .neg { color:var(--red); }
  .unpriced-note { color:var(--muted); font-size:12px; margin-bottom:6px; }
  .approx-note { color:var(--muted); font-size:11.5px; margin:-4px 0 16px; font-style:italic; }
  .charts { display:flex; gap:16px; flex-wrap:wrap; margin-bottom:22px; }
  .chart-card { background:var(--card); border:1px solid var(--border); border-radius:12px; padding:16px; flex:1; min-width:320px; max-width:480px; height:300px; }
  .toolbar { display:flex; gap:10px; flex-wrap:wrap; align-items:center; margin-bottom:12px; }
  .toolbar input, .toolbar select {
    background:var(--card); border:1px solid var(--border); color:var(--text);
    padding:8px 12px; border-radius:8px; font-size:13px;
  }
  .toolbar input { min-width:220px; }
  .count { color:var(--muted); font-size:12px; margin-left:auto; }
  table { border-collapse:collapse; width:100%; background:var(--card); border-radius:12px; overflow:hidden; font-size:12.5px; }
  th, td { padding:8px 11px; text-align:right; border-bottom:1px solid var(--border); white-space:nowrap; }
  th:nth-child(1), td:nth-child(1), th:nth-child(2), td:nth-child(2) { text-align:left; white-space:normal; }
  th { color:var(--muted); font-weight:500; text-transform:uppercase; font-size:10px; letter-spacing:.03em;
       cursor:pointer; user-select:none; position:sticky; top:0; background:var(--card); }
  th:hover { color:var(--text); }
  th.sorted-asc::after { content:" \\25B2"; font-size:9px; }
  th.sorted-desc::after { content:" \\25BC"; font-size:9px; }
  tr:hover td { background:var(--hover); }
  tr:last-child td { border-bottom:none; }
  .tag { font-size:10px; padding:2px 6px; border-radius:4px; background:#232838; color:var(--muted); }
  .tag.live { color:var(--green); background:#123822; }
  .ticker-link { cursor:pointer; text-decoration:underline dotted; text-underline-offset:2px; }
  .ticker-link:hover { color:var(--accent); }
  tr.total-row td { font-weight:700; border-top:2px solid var(--accent); background:#161c2c; }
  tr.row-link { cursor:pointer; }
  .table-wrap { max-height:560px; overflow:auto; border-radius:12px; }
  input.hscroll {
    -webkit-appearance:none; appearance:none; width:100%; height:6px; margin:10px 0 4px;
    background:var(--border); border-radius:4px; outline:none; cursor:pointer;
  }
  input.hscroll:disabled { opacity:.3; cursor:default; }
  input.hscroll::-webkit-slider-thumb {
    -webkit-appearance:none; appearance:none; width:16px; height:16px; border-radius:50%;
    background:var(--accent); cursor:pointer; border:2px solid var(--bg);
  }
  input.hscroll::-moz-range-thumb {
    width:16px; height:16px; border-radius:50%; background:var(--accent); cursor:pointer;
    border:2px solid var(--bg);
  }
  input.hscroll::-moz-range-track { background:var(--border); height:6px; border-radius:4px; }
</style>
</head>
<body>
  <h1>Portfolio Dashboard</h1>
  <div class="sub" id="generatedAt"></div>
  <div class="fxbar" id="fxBar"></div>
  <div class="tabs" id="tabs"></div>
  <div id="pages"></div>

<script>
const DATA = __DATA_JSON__;

const fmt = (n) => n === null || n === undefined ? '—' :
  n.toLocaleString('es-AR', {minimumFractionDigits:2, maximumFractionDigits:2});
const fmtPct = (n) => n === null || n === undefined ? '—' : (n>=0?'+':'') + n.toFixed(2) + '%';
const plClass = (n) => n === null || n === undefined ? '' : (n >= 0 ? 'pos' : 'neg');
// "en vivo" cubre tanto PPI como el fallback de Yahoo Finance (segunda
// fuente para instrumentos en USD que PPI no cotiza o no tiene datos ahora).
const isLiveSource = (s) => s === 'PPI (en vivo)' || s === 'Yahoo Finance';

document.getElementById('generatedAt').textContent = 'Generado ' + DATA.generated_at;
document.getElementById('fxBar').textContent = DATA.fx_rate
  ? `Tipo de cambio usado: 1 USD = ${fmt(DATA.fx_rate)} ARS (${DATA.fx_source})`
  : `Tipo de cambio no disponible (${DATA.fx_source}) -- las columnas convertidas quedan vacías`;

const MARKET_KEYS = ['general', 'usd', 'cedears', 'merval', 'rsu', 'bonds', 'tx'];
const TAB_LABELS = { general: 'General', usd: 'US Stocks', cedears: 'Cedears', merval: 'Acciones Merval', rsu: 'RSU', bonds: 'Bonos', tx: 'Transacciones' };

// portfolios "reales" (no general/tx) -- para armar los totales combinados
// de la pestana General sin repetir la lista en cada lugar.
const PORTFOLIO_KEYS = ['usd', 'cedears', 'merval', 'rsu', 'bonds'];

const tabsEl = document.getElementById('tabs');
const pagesEl = document.getElementById('pages');

// Cada tabla muestra UNA sola moneda a la vez (columnas base + las 10
// columnas "duales" resueltas segun el filtro de moneda de esa pestana).
const DEFAULT_CURRENCY = { general: 'usd', usd: 'usd', cedears: 'ars', merval: 'ars', rsu: 'usd', bonds: 'ars' };

const COLUMNS_BASE = [
  {key:'key', label:'Ticker'},
  {key:'name', label:'Nombre'},
  {key:'sector', label:'Sector'},
  {key:'instrument_type', label:'Tipo'},
  {key:'units', label:'Unid.', num:true},
  {key:'units_sold', label:'Unid. Vend.', num:true},
  {key:'trend_30d', label:'Tend. 30d %', num:true, pct:true},
];
const COLUMNS_DUAL = [
  {key:'pct_portfolio', label:'% Portfolio', num:true, pct:true, dual:true},
  {key:'avg_cost', label:'Prom. Compra', num:true, dual:true},
  {key:'price', label:'Precio Actual', num:true, dual:true},
  {key:'invested', label:'Invertido', num:true, dual:true},
  {key:'value', label:'Valor Actual', num:true, dual:true},
  {key:'pl_abs', label:'P&L $', num:true, pl:true, dual:true},
  {key:'pl_pct', label:'P&L %', num:true, pl:true, pct:true, dual:true},
  {key:'cost_of_sales', label:'Costo Ventas', num:true, dual:true},
  {key:'income_from_sales', label:'Ingreso Ventas', num:true, dual:true},
  {key:'realized_abs', label:'P&L Realiz. $', num:true, pl:true, dual:true},
  {key:'realized_pct', label:'P&L Realiz. %', num:true, pl:true, pct:true, dual:true},
];

// La pestana general muestra UNA fila por portfolio (USD/Cedears/Merval),
// no una fila por producto -- son los totales de cada tabla de arriba.
const GENERAL_COLUMNS = [
  {key:'label', label:'Portfolio'},
  {key:'count', label:'Instrumentos', num:true},
  {key:'pct_portfolio', label:'% Portfolio', num:true, pct:true},
  {key:'unpriced', label:'Sin Precio', num:true},
  {key:'invested', label:'Invertido', num:true, showCurrency:true},
  {key:'value', label:'Valor Actual', num:true, showCurrency:true},
  {key:'pl_abs', label:'P&L $', num:true, pl:true, showCurrency:true},
  {key:'pl_pct', label:'P&L %', num:true, pl:true, pct:true, showCurrency:true},
  {key:'realized', label:'P&L Realizado', num:true, pl:true, showCurrency:true},
];

// Pestana "Transacciones": una fila por operacion (BUY/SELL), no agregada.
const TX_COLUMNS = [
  {key:'ticker', label:'Ticker'},
  {key:'market', label:'Tipo'},
  {key:'op', label:'Operacion'},
  {key:'date', label:'Fecha'},
  {key:'units', label:'Unidades', num:true},
  // primero todo en ARS, despues todo en USD (mismo orden que las tablas
  // por tipo de instrumento).
  {key:'price_ars', label:'Precio Operacion (ARS)', num:true},
  {key:'amount_ars', label:'Monto (ARS)', num:true},
  {key:'current_price_ars', label:'Cotizacion Actual (ARS)', num:true},
  {key:'pl_pct_ars', label:'P&L % (ARS)', num:true, pct:true, pl:true},
  {key:'price_usd', label:'Precio Operacion (USD)', num:true},
  {key:'amount_usd', label:'Monto (USD)', num:true},
  {key:'current_price_usd', label:'Cotizacion Actual (USD)', num:true},
  {key:'pl_pct_usd', label:'P&L % (USD)', num:true, pct:true, pl:true},
];

function getCols(marketKey) {
  if (marketKey === 'general') return GENERAL_COLUMNS;
  if (marketKey === 'tx') return TX_COLUMNS;
  return [...COLUMNS_BASE, ...COLUMNS_DUAL];
}

// Filas de un portfolio tal como las ve la pestana general: completas, o
// solo las compradas en el anio elegido si ese filtro esta activo. Los KPIs
// de la general se recalculan desde aca (en vez de usar los precalculados de
// DATA.markets[m].kpis) para que respeten el filtro de anio.
function generalMarketRows(m) {
  const year = currentYear('general');
  const rows = DATA.markets[m].rows;
  return year ? rows.filter(r => matchesBuyYear(r, year)) : rows;
}

function _unpricedCount(rows) {
  // mismo criterio que _market_kpis() en Python: sin valor en ninguna moneda.
  return rows.filter(r => r.value_ars === null && r.value_usd === null).length;
}

function buildGeneralRows(currency) {
  const totalValue = computeGeneralTotals(currency).value;
  return PORTFOLIO_KEYS.map(m => {
    const rows = generalMarketRows(m);
    const k = computeMarketKpis(rows, currency);
    return {
      label: TAB_LABELS[m],
      market: m,
      count: rows.length,
      unpriced: _unpricedCount(rows),
      invested: k.invested, value: k.value,
      pct_portfolio: totalValue ? (k.value / totalValue * 100) : null,
      pl_abs: k.pl_abs,
      pl_pct: k.pl_pct, realized: k.realized,
    };
  });
}

// Totales combinados de los 3 portfolios, para las pills de arriba de la
// pestana general (mismo total que antes mostraba la fila "Total").
function computeGeneralTotals(currency) {
  const parts = PORTFOLIO_KEYS.map(m => computeMarketKpis(generalMarketRows(m), currency));
  const invested = parts.reduce((s, k) => s + (k.invested || 0), 0);
  const value = parts.reduce((s, k) => s + (k.value || 0), 0);
  const pl_abs = value - invested;
  const pl_pct = invested ? (pl_abs / invested * 100) : null;
  const realized = parts.reduce((s, k) => s + (k.realized || 0), 0);
  return { invested, value, pl_abs, pl_pct, realized };
}

function computeGeneralUnpriced() {
  return PORTFOLIO_KEYS.reduce((s, m) => s + _unpricedCount(generalMarketRows(m)), 0);
}

function _sumOrNone(rows, key) {
  let has = false, sum = 0;
  for (const r of rows) {
    const v = r[key];
    if (v !== null && v !== undefined) { has = true; sum += v; }
  }
  return has ? sum : null;
}

// Recalcula los KPIs de una pestana por tipo de instrumento a partir de
// cualquier subconjunto de filas (ej. lo que queda despues de buscar o
// filtrar por sector), asi las pills de arriba siempre reflejan lo que se
// esta viendo en la tabla de abajo -- igual logica que _market_kpis() en Python.
function computeMarketKpis(rows, currency) {
  const invested = _sumOrNone(rows, `invested_${currency}`);
  const value = _sumOrNone(rows, `value_${currency}`);
  const pl_abs = (value !== null && invested !== null) ? (value - invested) : null;
  const pl_pct = (pl_abs !== null && invested) ? (pl_abs / invested * 100) : null;
  const realized = _sumOrNone(rows, `realized_abs_${currency}`);
  return { invested, value, pl_abs, pl_pct, realized };
}

function cellValue(r, c, currency) {
  return c.dual ? r[`${c.key}_${currency}`] : r[c.key];
}

function buildKpiGroup(k) {
  const items = [
    ['Invertido', fmt(k.invested), ''],
    ['Valor Actual', fmt(k.value), ''],
    ['P&L No Realiz.', fmt(k.pl_abs), plClass(k.pl_abs)],
    ['P&L No Realiz. %', fmtPct(k.pl_pct), plClass(k.pl_pct)],
    ['P&L Realizado', fmt(k.realized), plClass(k.realized)],
  ];
  return `<div class="kpis">${items.map(([label,val,cls]) =>
    `<div class="kpi"><div class="label">${label}</div><div class="val ${cls}">${val}</div></div>`
  ).join('')}</div>`;
}

function renderKpis(marketKey, filteredRows) {
  const currency = tableState[marketKey].currency;
  const k = marketKey === 'general'
    ? computeGeneralTotals(currency)
    : computeMarketKpis(filteredRows || filterRows(marketKey), currency);
  const title = currency === 'ars' ? 'En Pesos (ARS)' : 'En Dolares (USD)';
  document.getElementById(`kpis-${marketKey}`).innerHTML =
    `<h3 class="kpi-title">${title}</h3>${buildKpiGroup(k)}`;
}

function uniqueSorted(rows, key) {
  return [...new Set(rows.map(r => r[key]).filter(v => v !== null && v !== undefined && v !== ''))].sort();
}

// Anios en los que se compro algo, del mas reciente al mas viejo. Una
// posicion puede tener compras en varios anios, asi que buy_years es una
// lista por fila y hay que aplanarla antes de deduplicar.
function uniqueYears(rows) {
  return [...new Set(rows.flatMap(r => r.buy_years || []))].sort().reverse();
}

// Anios con compras en cualquier portfolio: lo usan la pestana general y la
// de transacciones, que no corresponden a un solo tipo de instrumento.
function allBuyYears() {
  return uniqueYears(PORTFOLIO_KEYS.flatMap(m => DATA.markets[m].rows));
}

function yearSelectHtml(marketKey, years) {
  return `<select data-role="year" data-market="${marketKey}">
        <option value="">Todos los anios</option>
        ${years.map(y => `<option value="${y}">Comprado en ${y}</option>`).join('')}
      </select>`;
}

// Anio seleccionado en esa pestana ('' = sin filtrar).
function currentYear(marketKey) {
  const el = document.querySelector(`[data-role="year"][data-market="${marketKey}"]`);
  return el ? el.value : '';
}

function buildPage(marketKey) {
  const isGeneral = marketKey === 'general';
  const isTx = marketKey === 'tx';
  const defaultCurrency = DEFAULT_CURRENCY[marketKey];

  const page = document.createElement('div');
  page.className = 'page';
  page.id = 'page-' + marketKey;

  let html = '';

  if (isTx) {
    const typeOptions = PORTFOLIO_KEYS;
    html += `<div class="toolbar">
      <input type="text" placeholder="Buscar ticker..." data-role="search" data-market="${marketKey}">
      <select data-role="type" data-market="${marketKey}">
        <option value="">Todos los portfolios</option>
        ${typeOptions.map(t => `<option value="${t}">${TAB_LABELS[t]}</option>`).join('')}
      </select>
      ${yearSelectHtml(marketKey, allBuyYears())}
      <span class="count" data-role="count" data-market="${marketKey}"></span>
    </div>`;
  } else if (isGeneral) {
    const anyApprox = PORTFOLIO_KEYS.some(m => DATA.markets[m].rows.some(r => r.fx_approx));
    html += `<div id="kpis-${marketKey}"></div>`;
    const unpriced = computeGeneralUnpriced();
    if (unpriced > 0) {
      html += `<div class="unpriced-note">${unpriced} instrumento(s) sin precio actual disponible.</div>`;
    }
    if (anyApprox) {
      html += `<div class="approx-note">Uno o mas portfolios tienen posiciones con montos aproximados en la moneda no nativa (se convirtieron al tipo de cambio de HOY para esa operacion puntual, no historico). Mira el detalle en la pestana de cada tipo de instrumento.</div>`;
    }
    html += `<div class="toolbar">
      ${yearSelectHtml(marketKey, allBuyYears())}
      <select data-role="currency" data-market="${marketKey}">
        <option value="ars" ${defaultCurrency==='ars'?'selected':''}>Moneda: ARS</option>
        <option value="usd" ${defaultCurrency==='usd'?'selected':''}>Moneda: USD</option>
      </select>
    </div>`;
  } else {
    const rows = DATA.markets[marketKey].rows;
    const hasApprox = rows.some(r => r.fx_approx);
    html += `<div id="kpis-${marketKey}"></div>`;
    html += `<div class="unpriced-note">Instrumentos sin precio: ${DATA.markets[marketKey].kpis.unpriced}</div>`;
    if (hasApprox) {
      html += `<div class="approx-note">Algunas operaciones de esta posicion no tienen cargado el monto en la otra moneda (columna opcional en la hoja de transacciones), asi que para esas puntualmente se aproximo al tipo de cambio de HOY en vez del historico de esa operacion. El resto usa el monto real que cargaste.</div>`;
    }
    html += `<div class="charts">
      <div class="chart-card"><canvas id="bar-${marketKey}"></canvas></div>
      <div class="chart-card"><canvas id="pie-${marketKey}"></canvas></div>
    </div>`;

    const sectorOptions = uniqueSorted(rows, 'sector');
    const instTypeOptions = uniqueSorted(rows, 'instrument_type');
    html += `<div class="toolbar">
      <input type="text" placeholder="Buscar ticker o nombre..." data-role="search" data-market="${marketKey}">
      <select data-role="sector" data-market="${marketKey}">
        <option value="">Todos los sectores</option>
        ${sectorOptions.map(s => `<option value="${s}">${s}</option>`).join('')}
      </select>
      <select data-role="insttype" data-market="${marketKey}">
        <option value="">Todos los tipos</option>
        ${instTypeOptions.map(t => `<option value="${t}">${t}</option>`).join('')}
      </select>
      ${yearSelectHtml(marketKey, uniqueYears(rows))}
      <select data-role="currency" data-market="${marketKey}">
        <option value="ars" ${defaultCurrency==='ars'?'selected':''}>Moneda: ARS</option>
        <option value="usd" ${defaultCurrency==='usd'?'selected':''}>Moneda: USD</option>
      </select>
      <span class="count" data-role="count" data-market="${marketKey}"></span>
    </div>`;
  }

  html += `<div class="table-wrap" id="wrap-${marketKey}"><table id="table-${marketKey}">
    <thead><tr data-role="thead" data-market="${marketKey}"></tr></thead>
    <tbody></tbody>
  </table></div>
  <input type="range" class="hscroll" id="hscroll-${marketKey}" min="0" max="0" value="0" step="1">`;

  page.innerHTML = html;
  return page;
}

function renderHead(marketKey) {
  const currency = tableState[marketKey].currency;
  const cols = getCols(marketKey);
  const theadRow = document.querySelector(`thead tr[data-role="thead"][data-market="${marketKey}"]`);
  theadRow.innerHTML = cols.map(c => {
    const label = (c.dual || c.showCurrency) ? `${c.label} (${currency.toUpperCase()})` : c.label;
    return `<th data-key="${c.key}" data-market="${marketKey}">${label}</th>`;
  }).join('');
  wireSortHeaders(marketKey);
  const sort = tableState[marketKey].sort;
  const activeTh = theadRow.querySelector(`th[data-key="${sort.key}"]`);
  if (activeTh) activeTh.classList.add(sort.dir === 1 ? 'sorted-asc' : 'sorted-desc');
}

function wireSortHeaders(marketKey) {
  const cols = getCols(marketKey);
  document.querySelectorAll(`#table-${marketKey} th`).forEach(th => {
    th.addEventListener('click', () => {
      const key = th.dataset.key;
      const col = cols.find(c => c.key === key);
      const cur = tableState[marketKey].sort;
      const dir = (cur && cur.key === key) ? -cur.dir : -1;
      tableState[marketKey].sort = { key, dir, dual: !!(col && col.dual) };
      document.querySelectorAll(`#table-${marketKey} th`).forEach(h => h.classList.remove('sorted-asc','sorted-desc'));
      th.classList.add(dir === 1 ? 'sorted-asc' : 'sorted-desc');
      applyFilters(marketKey, cols);
    });
  });
}

const hscrollSync = {};

function wireHScroll(marketKey) {
  const wrap = document.getElementById(`wrap-${marketKey}`);
  const slider = document.getElementById(`hscroll-${marketKey}`);

  const syncMax = () => {
    // page might be display:none (inactive tab) -> scrollWidth/clientWidth
    // both read 0 then; harmless, we resync when the tab becomes active.
    const max = Math.max(0, wrap.scrollWidth - wrap.clientWidth);
    slider.max = max;
    slider.disabled = max <= 0;
    if (Number(slider.value) > max) slider.value = max;
  };

  hscrollSync[marketKey] = syncMax;
  syncMax();
  window.addEventListener('resize', syncMax);
  setTimeout(syncMax, 50);

  slider.addEventListener('input', () => { wrap.scrollLeft = Number(slider.value); });
  wrap.addEventListener('scroll', () => { slider.value = wrap.scrollLeft; }, { passive: true });
}

function renderRows(marketKey, rows, cols) {
  const currency = tableState[marketKey].currency;
  const tbody = document.querySelector(`#table-${marketKey} tbody`);
  tbody.innerHTML = rows.map(r => {
    const rowClass = r.market ? 'row-link' : '';
    const rowAttr = r.market ? ` data-market-link="${r.market}"` : '';
    return `<tr class="${rowClass}"${rowAttr}>${cols.map(c => {
    let v = cellValue(r, c, currency);
    if (c.key === 'key') {
      return `<td><strong class="ticker-link" data-ticker="${r.key}">${r.key}</strong> <span class="tag ${isLiveSource(r.price_source)?'live':''}">${r.price_source||''}</span></td>`;
    }
    if (c.key === 'ticker') {
      const src = r.current_price_source;
      const tag = src ? `<span class="tag ${isLiveSource(src)?'live':''}">${src}</span>` : '';
      return `<td><strong class="ticker-link" data-ticker="${v}">${v}</strong> ${tag}</td>`;
    }
    if (c.key === 'label') return `<td><strong>${v}</strong></td>`;
    if (c.key === 'market') return `<td>${TAB_LABELS[v] || v}</td>`;
    if (c.key === 'op') return `<td class="${v==='BUY'?'pos':(v==='SELL'?'neg':'')}">${v}</td>`;
    if (v === null || v === undefined) return `<td>—</td>`;
    if (c.pct) return `<td class="${c.pl?plClass(v):''}">${fmtPct(v)}</td>`;
    if (c.pl) return `<td class="${plClass(v)}">${fmt(v)}</td>`;
    if (c.num) return `<td>${fmt(v)}</td>`;
    return `<td>${v}</td>`;
    }).join('')}</tr>`;
  }).join('');
  const countEl = document.querySelector(`[data-role="count"][data-market="${marketKey}"]`);
  if (countEl) {
    const total = marketKey === 'tx' ? DATA.transactions.length : DATA.markets[marketKey].rows.length;
    countEl.textContent = `${rows.length} de ${total}`;
  }
}

// Filas que pasan los filtros activos (busqueda/sector/tipo) de esa
// pestana, sin ordenar todavia -- lo usan tanto la tabla como los graficos,
// asi ambos quedan siempre en sincro con lo que el usuario esta filtrando.
function filterRows(marketKey) {
  const currency = tableState[marketKey].currency;

  if (marketKey === 'general') {
    return buildGeneralRows(currency);
  }
  if (marketKey === 'tx') {
    const searchEl = document.querySelector(`[data-role="search"][data-market="tx"]`);
    const typeEl = document.querySelector(`[data-role="type"][data-market="tx"]`);
    const search = (searchEl.value || '').toLowerCase();
    const type = typeEl.value;
    // aca cada fila ES una operacion, asi que el filtro de anio se aplica
    // sobre la fecha de la operacion misma.
    const year = currentYear('tx');
    return DATA.transactions.filter(r => {
      if (search && !r.ticker.toLowerCase().includes(search)) return false;
      if (type && r.market !== type) return false;
      if (year && r.year !== year) return false;
      return true;
    });
  }
  const allRows = DATA.markets[marketKey].rows;
  const searchEl = document.querySelector(`[data-role="search"][data-market="${marketKey}"]`);
  const sectorEl = document.querySelector(`[data-role="sector"][data-market="${marketKey}"]`);
  const instTypeEl = document.querySelector(`[data-role="insttype"][data-market="${marketKey}"]`);
  const search = (searchEl.value || '').toLowerCase();
  const sector = sectorEl.value;
  const instType = instTypeEl ? instTypeEl.value : '';
  const year = currentYear(marketKey);
  return allRows.filter(r => {
    if (search && !(`${r.key} ${r.name}`.toLowerCase().includes(search))) return false;
    if (sector && r.sector !== sector) return false;
    if (instType && r.instrument_type !== instType) return false;
    if (year && !matchesBuyYear(r, year)) return false;
    return true;
  });
}

// Una posicion pasa el filtro si tuvo AL MENOS UNA compra en ese anio. Se
// muestra entera (todas las unidades y todo el P&L), no prorrateada al
// tramo comprado ese anio: el filtro responde "que compre en 2024", no
// "cuanto de lo que tengo hoy corresponde a 2024".
function matchesBuyYear(row, year) {
  return (row.buy_years || []).includes(year);
}

function applyFilters(marketKey, cols) {
  const currency = tableState[marketKey].currency;
  let rows = filterRows(marketKey);

  const sortState = tableState[marketKey].sort;
  if (sortState) {
    const { key, dir, dual } = sortState;
    const resolveKey = r => dual ? r[`${key}_${currency}`] : r[key];
    rows = [...rows].sort((a, b) => {
      let av = resolveKey(a), bv = resolveKey(b);
      const aNull = av === null || av === undefined || av === '';
      const bNull = bv === null || bv === undefined || bv === '';
      if (aNull && bNull) return 0;
      if (aNull) return 1;
      if (bNull) return -1;
      if (typeof av === 'string') return dir * av.localeCompare(bv);
      return dir * (av - bv);
    });
  }

  renderRows(marketKey, rows, cols);
  // los graficos y las pills de KPIs (no existen en Transacciones) se
  // recalculan con el mismo subconjunto filtrado que se ve en la tabla.
  if (marketKey !== 'general' && marketKey !== 'tx') renderCharts(marketKey, rows);
  if (marketKey !== 'tx') renderKpis(marketKey, rows);
}

const tableState = {};
const chartInstances = {};

function renderCharts(marketKey, filteredRows) {
  // General (tabla resumen de 3 filas) y Transacciones no tienen graficos.
  if (marketKey === 'general' || marketKey === 'tx') return;

  const currency = tableState[marketKey].currency;
  const cur = currency.toUpperCase();
  if (chartInstances[marketKey]) {
    chartInstances[marketKey].forEach(c => { if (c && typeof c.destroy === 'function') c.destroy(); });
  }

  // si no se paso un subconjunto ya filtrado, usar todo el portfolio
  // (caso init, antes de que el usuario aplique ningun filtro).
  const rows = filteredRows || DATA.markets[marketKey].rows;
  const top = [...rows].sort((a,b) => (b[`value_${currency}`]||0)-(a[`value_${currency}`]||0)).slice(0, 10);
  const bar = new Chart(document.getElementById(`bar-${marketKey}`), {
    type: 'bar',
    data: { labels: top.map(r=>r.key), datasets: [
      {label:`Invertido (${cur})`, data: top.map(r=>r[`invested_${currency}`]), backgroundColor:'#6ea8fe'},
      {label:`Valor Actual (${cur})`, data: top.map(r=>r[`value_${currency}`]||0), backgroundColor:'#3ddc84'},
    ]},
    options: { plugins:{legend:{labels:{color:'#e7e9ee'}}}, scales:{
      x:{ticks:{color:'#9aa2b1'}}, y:{ticks:{color:'#9aa2b1'}}
    }}
  });
  const palette = ['#6ea8fe','#3ddc84','#ffb84d','#ff5c5c','#c792ea','#4dd0e1','#f78fb3','#a0c980','#ffd166','#8d99ae'];
  const pie = new Chart(document.getElementById(`pie-${marketKey}`), {
    type: 'doughnut',
    data: { labels: top.map(r=>r.key), datasets: [{ data: top.map(r=>r[`value_${currency}`]||0), backgroundColor: palette }] },
    options: { plugins:{legend:{position:'bottom', labels:{color:'#e7e9ee', boxWidth:10, font:{size:10}}}} }
  });
  chartInstances[marketKey] = [bar, pie];
}

function initPage(marketKey) {
  const isGeneral = marketKey === 'general';
  const isTx = marketKey === 'tx';
  const cols = getCols(marketKey);
  tableState[marketKey] = {
    currency: DEFAULT_CURRENCY[marketKey] || 'ars',
    sort: {
      key: isGeneral ? 'pl_abs' : (isTx ? 'date' : 'value'),
      dir: -1,
      dual: !isGeneral && !isTx,
    },
  };

  renderHead(marketKey);
  applyFilters(marketKey, cols); // tambien dispara renderCharts y renderKpis con el set filtrado
  wireHScroll(marketKey);

  if (isTx) {
    document.querySelector(`[data-role="search"][data-market="tx"]`)
      .addEventListener('input', () => applyFilters('tx', cols));
    document.querySelector(`[data-role="type"][data-market="tx"]`)
      .addEventListener('change', () => applyFilters('tx', cols));
  } else if (!isGeneral) {
    document.querySelector(`[data-role="search"][data-market="${marketKey}"]`)
      .addEventListener('input', () => applyFilters(marketKey, cols));
    document.querySelector(`[data-role="sector"][data-market="${marketKey}"]`)
      .addEventListener('change', () => applyFilters(marketKey, cols));
    document.querySelector(`[data-role="insttype"][data-market="${marketKey}"]`)
      .addEventListener('change', () => applyFilters(marketKey, cols));
  }

  // el filtro de anio existe en las tres clases de pestana (general, por tipo
  // de instrumento y transacciones), asi que se cablea una sola vez aca.
  const yearEl = document.querySelector(`[data-role="year"][data-market="${marketKey}"]`);
  if (yearEl) {
    yearEl.addEventListener('change', () => applyFilters(marketKey, cols));
  }

  const currencyEl = document.querySelector(`[data-role="currency"][data-market="${marketKey}"]`);
  if (currencyEl) {
    currencyEl.addEventListener('change', (e) => {
      tableState[marketKey].currency = e.target.value;
      renderHead(marketKey);
      applyFilters(marketKey, cols); // tambien dispara renderCharts y renderKpis con el set filtrado
      if (hscrollSync[marketKey]) setTimeout(hscrollSync[marketKey], 0);
    });
  }
}

// Click any ticker (en las tablas por tipo o en general) -> salta a la
// pestana de Transacciones filtrada por ese ticker.
function goToTicker(ticker) {
  selectTab('tx');
  const searchEl = document.querySelector(`[data-role="search"][data-market="tx"]`);
  searchEl.value = ticker;
  applyFilters('tx', getCols('tx'));
}

pagesEl.addEventListener('click', (e) => {
  const t = e.target.closest('.ticker-link');
  if (t) { goToTicker(t.dataset.ticker); return; }

  // click en una fila de la pestana general (no el Total) -> ir al detalle
  const row = e.target.closest('tr[data-market-link]');
  if (row) { selectTab(row.dataset.marketLink); }
});

const tabEls = {};

function selectTab(key) {
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  tabEls[key].classList.add('active');
  document.getElementById('page-'+key).classList.add('active');
  // the page was display:none until now, so its scrollWidth/clientWidth
  // were unreadable -- recompute the slider range now that it's visible.
  if (hscrollSync[key]) setTimeout(hscrollSync[key], 0);
}

MARKET_KEYS.forEach((key, i) => {
  const tab = document.createElement('div');
  tab.className = 'tab' + (i===0 ? ' active' : '');
  tab.textContent = TAB_LABELS[key];
  tab.addEventListener('click', () => selectTab(key));
  tabsEl.appendChild(tab);
  tabEls[key] = tab;

  const page = buildPage(key);
  if (i===0) page.classList.add('active');
  pagesEl.appendChild(page);
});

MARKET_KEYS.forEach(initPage);
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path", help="Portfolio - Transacciones.xlsx")
    ap.add_argument("--out-html", default="Portfolio Dashboard.html")
    ap.add_argument("--out-xlsx", default=None,
                     help="Si se pasa, ademas genera un xlsx con la misma info (opcional).")
    args = ap.parse_args()

    reports, fx_rate, fx_source, transactions = build_all_reports(args.xlsx_path)

    render_html(reports, fx_rate, fx_source, transactions, args.out_html)

    if args.out_xlsx:
        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        write_dashboard_sheet(out_wb, "dashboard-usd", reports["usd"])
        write_portfolio_sheet(out_wb, "portfolio-usd", reports["usd"])
        write_dashboard_sheet(out_wb, "dashboard-cedears", reports["cedears"])
        write_portfolio_sheet(out_wb, "portfolio-cedears", reports["cedears"])
        write_dashboard_sheet(out_wb, "dashboard-merval", reports["merval"])
        write_portfolio_sheet(out_wb, "portfolio-merval", reports["merval"])
        write_dashboard_sheet(out_wb, "dashboard-rsu", reports["rsu"])
        write_portfolio_sheet(out_wb, "portfolio-rsu", reports["rsu"])
        write_dashboard_sheet(out_wb, "dashboard-bonds", reports["bonds"])
        write_portfolio_sheet(out_wb, "portfolio-bonds", reports["bonds"])
        write_general_dashboard(out_wb, reports)
        out_wb.move_sheet("dashboard-general", offset=-11)
        out_wb.save(args.out_xlsx)

    total_rows = sum(len(r) for r in reports.values())
    unpriced_rows = [r for rs in reports.values() for r in rs
                      if r["value_ars"] is None and r["value_usd"] is None]
    print(f"Listo: {args.out_html}")
    if args.out_xlsx:
        print(f"Listo: {args.out_xlsx}")
    print(f"  {total_rows} posiciones procesadas, {len(unpriced_rows)} sin precio disponible.")
    if unpriced_rows:
        detalle = ", ".join(f"{r['key']} ({MARKET_LABELS[r['market']]})" for r in unpriced_rows)
        print(f"    Sin precio: {detalle}")
        print(f"    -> Cargales un precio en la columna \"Precio Manual\" de la hoja "
              f"instrumentos, o revisa que el Key coincida exactamente con el ticker "
              f"real (PPI y/o Yahoo Finance) para que puedan cotizar solos.")
        for r in unpriced_rows:
            if r.get("price_debug_note"):
                print(f"      - {r['key']}: {r['price_debug_note']}")
    print(f"  Tipo de cambio: {fx_rate if fx_rate else 'N/D'} ({fx_source})")
    if not fx_rate:
        print(f"    -> Ni PPI (dolar MEP AL30/AL30D) ni la hoja config tienen un tipo de "
              f"cambio disponible. Revisa tus credenciales PPI_PUBLIC_KEY/PPI_PRIVATE_KEY, "
              f"o cargue un valor en config!B2 (\"USD/ARS Manual\") como respaldo. Mientras "
              f"tanto, las columnas convertidas a la otra moneda quedan vacias.")


if __name__ == "__main__":
    main()
